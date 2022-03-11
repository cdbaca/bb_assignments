import openpyxl
import json

# PURPOSE:
# 1. get course data/parentId of course content folder,
# 2. put in json format necessary for the Bb API assignments POST path

# BEFORE STARTING:
# Use pgAdmin to get correct data: "Find Course Contents Folder in Courses" query
# Save query as 'course_data.xlsx' in bb_assignments folder

# CURRENTLY LOADED: 3 Summer 2022 courses

assignment_workbook = openpyxl.load_workbook('data/course_data.xlsx', read_only=False)
assignment_data = assignment_workbook['Sheet1']

course_id = ''
parent_id = ''

course_dict = {}
course_parent_data = {}

# check Bb API docs if the json doesn't parse properly went sent in the POST request. Some fields were removed.
base_dict = """{
  "parentId": "_1348966_1",
  "title": "test assignment number 2",
  "instructions": "<h4>Test Instructions for the Assignment</h4>",
  "description": "test description",
  "position": 0,
  "availability": {
    "available": "Yes",
    "allowGuests": true,
    "adaptiveRelease": {
      "start": "2022-02-07T17:07:19.365Z",
      "end": "2022-02-07T17:07:19.365Z"
    }
  },
  "grading": {
    "due": "2022-02-07T17:07:19.365Z",
    "attemptsAllowed": 0,
    "isUnlimitedAttemptsAllowed": true
  },
  "score": {
    "possible": 0
  }
}"""

# create baseline json files

base_dict = json.loads(base_dict)

for row_num in range(2, assignment_data.max_row+1):
    course_id = assignment_data.cell(row=row_num, column=4).value
    parent_id = assignment_data.cell(row=row_num, column=3).value
    course_parent_data[course_id] = parent_id

for k,v in course_parent_data.items():
    course_dict[k] = base_dict

# for some reason, I need to place the dicts in a json file before dealing with them,
# otherwise for loop at the bottom won't work properly
# maybe it's because of the variables I'm using?

f = open("data/course_parent_template.json", "w")
json.dump(course_parent_data, f)
f.close()

f = open("data/course_dict.json", "w")
json.dump(course_dict, f)
f.close()

with open("data/course_parent_template.json", "r") as read_file:
    dict1 = json.load(read_file)

with open("data/course_dict.json", "r") as read_file:
    dict2 = json.load(read_file)

# put json data into new json file: bb_course_data, to be used in add_assign.py

for k in dict1:
  if k in dict2:
    dict2[k]["parentId"] = dict1[k]

f = open("data/bb_course_data.json", "w")
json.dump(dict2, f)
f.close()

print(json.dumps(dict2,indent=4, separators=(',', ': ')))